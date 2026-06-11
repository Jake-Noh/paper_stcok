import io
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from data.db import get_stock_results, get_products
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ── 최신 운영재고 목표 로드 ────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def _load_latest_targets():
    rows = get_stock_results(limit=500)
    if not rows:
        return {}, ""
    latest_calc = max(r["calc_yyyymm"] for r in rows)
    all_latest  = [r for r in rows if r["calc_yyyymm"] == latest_calc]
    min_target  = min(r.get("target_yyyymm", "999999") for r in all_latest)
    target_rows = [r for r in all_latest if r.get("target_yyyymm") == min_target]
    return {r["product_id"]: r for r in target_rows}, min_target


# ── 현재고 입력 양식 생성 ────────────────────────────────────────────────────
def _make_stock_template(products, targets):
    rows = []
    for p in products:
        pid    = p["product_id"]
        pname  = p["product_name"]
        t      = targets.get(pid, {})
        rows.append({
            "product_id":       pid,
            "product_name":     pname,
            "current_stock(톤)": "",          # ← 입력 칸
            "운영재고목표(참고)": round(t.get("operating_stock", 0), 1),
            "안전재고(참고)":    round(t.get("safety_stock_independent", 0), 1),
        })
    df = pd.DataFrame(rows)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="현재고입력")
        ws = writer.sheets["현재고입력"]

        thin = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        hdr_fill  = PatternFill("solid", fgColor="1F5C99")
        hdr_font  = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
        body_font = Font(name="맑은 고딕", size=9)
        input_fill = PatternFill("solid", fgColor="FFF9C4")   # 노랑 — 입력 칸
        ref_fill   = PatternFill("solid", fgColor="F5F5F5")   # 회색 — 참고용
        center     = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = thin

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.font      = body_font
                cell.alignment = center
                cell.border    = thin
                cell.fill = input_fill if col_idx == 3 else ref_fill

        col_widths = [12, 16, 18, 18, 14]
        for i, w in enumerate(col_widths, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 17

        ws.freeze_panes = "A2"

        # 범례
        ws["G2"] = "※ 노란색 칸(current_stock)에만 수치를 입력하세요."
        ws["G2"].font = Font(name="맑은 고딕", size=9, color="856404")

    buf.seek(0)
    return buf


# ── 계산 로직 ─────────────────────────────────────────────────────────────────
def _calc_orders(cur_stocks: dict, targets: dict, name_map: dict) -> list:
    rows = []
    for pid, t in targets.items():
        cur        = float(cur_stocks.get(pid, 0.0))
        op_target  = float(t.get("operating_stock", 0.0))
        safety     = float(t.get("safety_stock_independent", 0.0))
        d_prime    = float(t.get("d_prime", 1.0)) or 1.0
        avg_lt     = float(t.get("avg_lt", 5.0))
        target_days= float(t.get("operating_days", 0.0))

        need      = op_target - cur
        order_qty = max(0.0, need)
        surplus   = max(0.0, -need)
        cur_days  = cur / d_prime

        if cur < safety:
            status      = "🔴 위험"
            status_note = "안전재고 미달 — 즉시 발주"
        elif cur < op_target:
            status      = "🟡 주의"
            status_note = f"목표까지 {need:,.0f}톤 부족"
        else:
            status      = "🟢 정상"
            status_note = f"목표 대비 {surplus:,.0f}톤 초과"

        lt_cover      = d_prime * avg_lt
        order_timing  = "즉시 발주" if cur <= lt_cover else f"{int((cur - lt_cover) / d_prime)}일 후"

        rows.append({
            "지종":           name_map.get(pid, pid),
            "현재고(톤)":     round(cur, 1),
            "현재고(일)":     round(cur_days, 1),
            "목표재고(톤)":   round(op_target, 1),
            "목표재고(일)":   round(target_days, 1),
            "안전재고(톤)":   round(safety, 1),
            "발주필요량(톤)": round(order_qty, 1),
            "초과재고(톤)":   round(surplus, 1),
            "상태":           status,
            "발주시점":       order_timing,
            "비고":           status_note,
            "_pid":           pid,
            "_cur":           cur,
            "_op_target":     op_target,
            "_safety":        safety,
            "_order_qty":     order_qty,
        })

    rows.sort(key=lambda r: {"🔴 위험": 0, "🟡 주의": 1, "🟢 정상": 2}.get(r["상태"], 9))
    return rows


# ── 결과 출력 ─────────────────────────────────────────────────────────────────
def _render_results(rows: list, target_month: str):
    danger_cnt  = sum(1 for r in rows if r["상태"] == "🔴 위험")
    caution_cnt = sum(1 for r in rows if r["상태"] == "🟡 주의")
    ok_cnt      = sum(1 for r in rows if r["상태"] == "🟢 정상")
    total_order = sum(r["_order_qty"] for r in rows)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("즉시 발주 필요",
              f"{danger_cnt + caution_cnt}개 지종",
              delta=f"위험 {danger_cnt}건" if danger_cnt else None,
              delta_color="inverse" if danger_cnt else "off")
    k2.metric("총 발주 필요량", f"{total_order:,.0f} 톤")
    k3.metric("정상 지종",      f"{ok_cnt}개 지종")
    k4.metric("전체 지종",      f"{len(rows)}개")

    st.divider()

    # 상세 테이블
    st.subheader("📋 발주 필요량 상세")
    disp_cols = ["지종", "현재고(톤)", "현재고(일)", "목표재고(톤)", "목표재고(일)",
                 "안전재고(톤)", "발주필요량(톤)", "초과재고(톤)", "상태", "발주시점", "비고"]
    st.dataframe(pd.DataFrame(rows)[disp_cols], use_container_width=True, hide_index=True)

    # 위험·주의 강조
    urgent = [r for r in rows if r["상태"] in ("🔴 위험", "🟡 주의")]
    if urgent:
        st.error(f"**즉시 조치 필요 {len(urgent)}개 지종**")
        for r in urgent:
            st.markdown(
                f"{r['상태']} **{r['지종']}** — "
                f"현재고 {r['현재고(톤)']}톤 ({r['현재고(일)']}일) │ "
                f"발주 필요량 **{r['발주필요량(톤)']}톤** │ {r['비고']}"
            )

    st.divider()

    # 차트 1 : 현재고 vs 목표 vs 안전재고
    st.subheader("📊 현재고 vs 목표재고 vs 안전재고")
    pnames      = [r["지종"]       for r in rows]
    cur_vals    = [r["_cur"]       for r in rows]
    target_vals = [r["_op_target"] for r in rows]
    safety_vals = [r["_safety"]    for r in rows]

    fig = go.Figure()
    fig.add_bar(name="운영재고 목표", x=pnames, y=target_vals,
                marker_color="rgba(69,183,209,0.30)")
    fig.add_bar(name="현재고",        x=pnames, y=cur_vals,
                marker_color="#4ECDC4")
    fig.add_scatter(
        name="안전재고 기준선",
        x=pnames, y=safety_vals,
        mode="markers+lines",
        marker=dict(symbol="line-ew", size=14, color="#FF6B6B",
                    line=dict(width=2.5, color="#FF6B6B")),
        line=dict(color="#FF6B6B", width=1.5, dash="dot"),
    )
    fig.update_layout(barmode="overlay", height=400,
                      xaxis_title="지종", yaxis_title="재고량 (톤)",
                      legend=dict(orientation="h", yanchor="bottom", y=1.02),
                      hovermode="x unified")
    st.plotly_chart(fig, use_container_width=True)

    # 차트 2 : 발주 필요량
    order_rows = [r for r in rows if r["_order_qty"] > 0]
    if order_rows:
        st.subheader("📦 발주 필요량 (지종별)")
        o_names  = [r["지종"]        for r in order_rows]
        o_vals   = [r["_order_qty"]  for r in order_rows]
        o_colors = ["#FF6B6B" if r["상태"] == "🔴 위험" else "#FFB347" for r in order_rows]
        fig2 = go.Figure()
        fig2.add_bar(x=o_names, y=o_vals, marker_color=o_colors,
                     text=[f"{v:,.0f}톤" for v in o_vals], textposition="outside")
        fig2.update_layout(height=320, xaxis_title="지종",
                           yaxis_title="발주 필요량 (톤)", showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.success("✅ 모든 지종이 운영재고 목표를 충족합니다. 현재 발주 필요량 없음.")

    st.divider()

    # 엑셀 다운로드
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(rows)[disp_cols].to_excel(writer, index=False, sheet_name="발주필요량")
        ws = writer.sheets["발주필요량"]
        hdr_fill = PatternFill("solid", fgColor="1F5C99")
        hdr_font = Font(color="FFFFFF", bold=True, name="맑은 고딕", size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[col[0].column_letter].width = 15
    buf.seek(0)
    st.download_button(
        "📥 발주 필요량 엑셀 다운로드",
        data=buf,
        file_name=f"발주필요량_{target_month}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ── 메인 렌더 ─────────────────────────────────────────────────────────────────
def render_order_page():
    st.title("🚚 현재고 입력 · 발주 필요량 산출")
    st.caption("현재 창고 재고를 입력하면 운영재고 목표와 비교해 즉시 발주 필요량을 계산합니다.")

    products    = get_products()
    name_map    = {p["product_id"]: p["product_name"] for p in products}
    targets, target_month = _load_latest_targets()

    if not targets:
        st.warning("운영재고 산출 결과가 없습니다. '월별 실적 입력' 메뉴에서 먼저 산출을 실행하세요.")
        return

    st.info(
        f"운영재고 목표 기준월: **{target_month[:4]}년 {target_month[4:]}월**  |  "
        "아래에서 엑셀 업로드 또는 직접 입력 후 계산 버튼을 누르세요."
    )

    st.divider()

    # ── 탭: 엑셀 업로드 / 직접 입력 ─────────────────────────────────────────
    tab_xl, tab_manual = st.tabs(["📤 엑셀 업로드", "⌨️ 직접 입력"])

    cur_stocks = {}   # {pid: float}

    # ── 탭 1: 엑셀 업로드 ────────────────────────────────────────────────────
    with tab_xl:
        dl_col, up_col = st.columns(2, gap="large")

        with dl_col:
            st.markdown("##### 1. 양식 다운로드")
            st.caption("노란색 `current_stock(톤)` 칸에 현재고를 입력하세요.")
            st.download_button(
                "📥 현재고 입력양식 다운로드 (.xlsx)",
                data=_make_stock_template(products, targets),
                file_name=f"현재고입력양식_{target_month}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with up_col:
            st.markdown("##### 2. 작성 후 업로드")
            st.caption("필수 컬럼: `product_id` · `current_stock(톤)`")
            uploaded = st.file_uploader(
                "현재고 엑셀 파일 업로드",
                type=["xlsx", "xls"],
                key="stock_uploader",
                label_visibility="collapsed",
            )

        if uploaded:
            try:
                xl     = pd.ExcelFile(uploaded)
                sheet  = xl.sheet_names[0]
                df     = xl.parse(sheet)
                df.columns = [c.replace("(톤)", "").strip() for c in df.columns]

                if "product_id" not in df.columns or "current_stock" not in df.columns:
                    st.error(f"❌ 필수 컬럼 누락 — 필요: product_id, current_stock(톤) | 현재: {list(df.columns)}")
                else:
                    df["current_stock"] = pd.to_numeric(df["current_stock"], errors="coerce").fillna(0.0)
                    xl_stocks = {
                        str(row["product_id"]).strip(): float(row["current_stock"])
                        for _, row in df.iterrows()
                        if str(row["product_id"]).strip() in targets
                    }

                    st.markdown("**📋 업로드 확인**")
                    preview = df[["product_id", "current_stock"]].copy()
                    preview.columns = ["지종 ID", "현재고(톤)"]
                    st.dataframe(preview, use_container_width=True, hide_index=True)

                    if st.button("🔢 업로드 데이터로 발주 필요량 계산",
                                 type="primary", use_container_width=True, key="calc_xl"):
                        st.session_state["order_stocks"] = xl_stocks
                        st.session_state["order_source"] = "xl"
                        st.rerun()

            except Exception as e:
                st.error(f"파일 처리 오류: {e}")
                import traceback; st.code(traceback.format_exc())

    # ── 탭 2: 직접 입력 ──────────────────────────────────────────────────────
    with tab_manual:
        st.markdown("##### 지종별 현재고 입력 (톤)")
        cols_per_row = 4
        prod_list    = list(targets.keys())
        manual_vals  = {}

        for row_start in range(0, len(prod_list), cols_per_row):
            cols = st.columns(cols_per_row)
            for i, pid in enumerate(prod_list[row_start: row_start + cols_per_row]):
                pname      = name_map.get(pid, pid)
                op_target  = targets[pid].get("operating_stock", 0)
                saved_val  = st.session_state.get(f"mni_{pid}", 0.0)
                with cols[i]:
                    val = st.number_input(
                        f"{pname}",
                        min_value=0.0,
                        value=float(saved_val),
                        step=10.0,
                        format="%.1f",
                        help=f"목표: {op_target:,.0f}톤",
                        key=f"ni_{pid}",
                    )
                    manual_vals[pid] = val

        if st.button("🔢 입력값으로 발주 필요량 계산",
                     type="primary", use_container_width=True, key="calc_manual"):
            for pid, v in manual_vals.items():
                st.session_state[f"mni_{pid}"] = v
            st.session_state["order_stocks"] = dict(manual_vals)
            st.session_state["order_source"] = "manual"
            st.rerun()

    # ── 결과 표시 ─────────────────────────────────────────────────────────────
    if "order_stocks" in st.session_state:
        cur_stocks = st.session_state["order_stocks"]
        src_label  = "엑셀 업로드" if st.session_state.get("order_source") == "xl" else "직접 입력"
        st.divider()
        st.subheader(f"📊 발주 필요량 산출 결과  ·  입력 방식: {src_label}")

        rows = _calc_orders(cur_stocks, targets, name_map)
        _render_results(rows, target_month)
