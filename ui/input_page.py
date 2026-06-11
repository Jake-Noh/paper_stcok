import streamlit as st
import pandas as pd
import io
from datetime import datetime
from data.db import get_products, insert_monthly_sales, insert_leadtime_record, get_setting
from utils.validator import validate_sales_input, validate_leadtime_input


def _get_target_yyyymm():
    now = datetime.now()
    if now.month == 12:
        return f"{now.year + 1}01"
    return f"{now.year}{now.month + 1:02d}"


def _months_from_to(start_yyyymm, end_yyyymm):
    """start_yyyymm 부터 end_yyyymm 까지 월 리스트 반환."""
    months = []
    y, m = int(start_yyyymm[:4]), int(start_yyyymm[4:])
    ey, em = int(end_yyyymm[:4]), int(end_yyyymm[4:])
    while (y, m) <= (ey, em):
        months.append(f"{y}{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def _run_full_pipeline(target_yyyymm, products):
    from core.statistics_engine import StatisticsEngine
    from core.operating_stock import OperatingStockEngine
    from core.rules import RuleEngine
    from core.safety_stock import Z_TABLE
    from data.db import save_stock_result, get_monthly_sales
    from ml.demand_forecast import DemandForecaster

    stats_engine = StatisticsEngine()
    op_engine    = OperatingStockEngine()
    rule_engine  = RuleEngine()
    forecaster   = DemandForecaster()
    working_days = int(get_setting("working_days", "22"))

    # 대상월 ~ 해당 연도 12월까지 전체 산출
    target_year   = target_yyyymm[:4]
    end_yyyymm    = f"{target_year}12"
    month_list    = _months_from_to(target_yyyymm, end_yyyymm)
    total_steps   = len(products) * len(month_list)
    step          = 0
    progress_bar  = st.progress(0, text="운영재고 산출 시작...")

    # 지종별 통계는 한 번만 계산
    stats_cache = {}
    for p in products:
        pid = p["product_id"]
        sigma_d          = stats_engine.calc_sigma_d(pid)
        avg_lt, sigma_lt = stats_engine.calc_sigma_lt(pid)
        service_level    = p.get("service_level", 0.95)
        z                = Z_TABLE.get(service_level, 1.645)
        forecaster.fit(pid)
        stats_cache[pid] = {
            "sigma_d": sigma_d, "sigma_lt": sigma_lt,
            "avg_lt": avg_lt, "z": z,
            "pc": p["pc_days"], "service_level": service_level,
        }

    # 대상월 결과 (session_state 반환용)
    target_results = []

    for ym in month_list:
        ym_label = f"{ym[:4]}년 {ym[4:]}월"
        for p in products:
            pid = p["product_id"]
            step += 1
            progress_bar.progress(
                step / total_steps,
                text=f"[{p.get('product_name', pid)}] {ym_label} 계산 중... ({step}/{total_steps})"
            )

            sc = stats_cache[pid]

            # plan_qty 결정:
            #  1) 해당 월 실적/계획 DB에 있으면 우선 사용
            #  2) session_state에 수동 입력값 있으면 사용 (대상월만)
            #  3) 없으면 수요예측 자동 추정
            db_sales = get_monthly_sales(pid, 36)
            plan_qty = None
            for row in db_sales:
                if str(row.get("yyyymm", "")) == ym:
                    plan_qty = row.get("plan_qty") or row.get("actual_qty")
                    break

            if plan_qty is None and ym == target_yyyymm:
                plan_qty = st.session_state.get(f"plan_{pid}", None)

            if plan_qty is None:
                # 수요예측으로 자동 추정
                fc = forecaster.predict_next_month(pid)
                plan_qty = fc.get("recommended", 3000.0)

            stats = {
                "sigma_d":    sc["sigma_d"],
                "sigma_lt":   sc["sigma_lt"],
                "avg_lt":     sc["avg_lt"],
                "z":          sc["z"],
                "pc":         sc["pc"],
            }
            result = op_engine.build_full_result(
                pid, ym, sc["service_level"], stats, plan_qty, working_days
            )
            result = rule_engine.apply_all_rules(pid, result)
            result["product_name"] = p.get("product_name", pid)
            result["trend"]        = p.get("trend", "stable")
            save_stock_result(result)

            if ym == target_yyyymm:
                target_results.append(result)

    progress_bar.progress(1.0, text=f"완료! ({len(month_list)}개월 × {len(products)}개 지종)")
    return target_results, month_list


def _make_sales_template(products):
    rows = [
        {
            "product_id": p["product_id"],
            "product_name": p["product_name"],
            "yyyymm": "",
            "plan_qty(톤)": "",
            "actual_qty(톤)": "",
        }
        for p in products
    ]
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="판매실적")
        ws = writer.sheets["판매실적"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        # 헤더 색상
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    buf.seek(0)
    return buf


def _make_leadtime_template(products):
    rows = [
        {
            "product_id": p["product_id"],
            "product_name": p["product_name"],
            "order_date": "2026-01-01",
            "receipt_date": "2026-01-05",
            "lt_days(일)": "",
            "weight_qty(톤)": "",
        }
        for p in products
    ]
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="리드타임실적")
        ws = writer.sheets["리드타임실적"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill("solid", fgColor="70AD47")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    buf.seek(0)
    return buf


def _render_sales_upload_section(products, target_yyyymm):
    """판매 실적 업로드 전용 섹션"""

    st.markdown("##### 📥 양식 다운로드")
    st.download_button(
        label="판매실적 입력양식 다운로드 (.xlsx)",
        data=_make_sales_template(products),
        file_name=f"판매실적_입력양식_{target_yyyymm}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("##### 📤 파일 업로드")
    st.caption("필수 컬럼: `product_id` · `yyyymm` · `plan_qty(톤)` · `actual_qty(톤)`")
    uploaded = st.file_uploader(
        "판매 실적 엑셀 파일을 여기에 올려주세요",
        type=["xlsx", "xls"],
        key="sales_file_uploader",
        label_visibility="collapsed",
    )

    if uploaded is None:
        st.info("양식을 다운받아 작성한 뒤 업로드하세요.")
        return

    try:
        xl = pd.ExcelFile(uploaded)
        # "판매" 관련 시트가 있으면 우선 사용, 없으면 첫 번째 시트
        sales_sheets = [s for s in xl.sheet_names if "판매" in s or "sales" in s.lower()]
        sheet = sales_sheets[0] if sales_sheets else xl.sheet_names[0]
        df = xl.parse(sheet)
        # 컬럼명 정규화 (괄호 포함 버전 대응)
        df.columns = [c.replace("(톤)", "").strip() for c in df.columns]

        required = {"product_id", "plan_qty"}
        missing = required - set(df.columns)
        if missing:
            st.error(f"❌ 필수 컬럼 누락: {missing}  (읽은 시트: '{sheet}', 전체 시트: {xl.sheet_names})")
            return

        if "yyyymm" not in df.columns or df["yyyymm"].isnull().all():
            df["yyyymm"] = target_yyyymm
        else:
            df["yyyymm"] = df["yyyymm"].fillna(target_yyyymm).astype(str).str.replace("-", "").str[:6]

        if "actual_qty" not in df.columns:
            df["actual_qty"] = None

        # 미리보기
        st.markdown("**📋 업로드 데이터 미리보기**")
        preview_cols = [c for c in ["product_id", "yyyymm", "plan_qty", "actual_qty"] if c in df.columns]
        st.dataframe(df[preview_cols], use_container_width=True, hide_index=True)
        st.caption(f"총 {len(df)}행")

        if st.button("💾 판매 실적 저장", type="primary", use_container_width=True, key="save_sales_btn"):
            success, fail, err_msgs = 0, 0, []
            for _, row in df.iterrows():
                pid = str(row["product_id"]).strip()
                ym = str(row["yyyymm"]).strip()
                plan = float(row["plan_qty"]) if pd.notna(row["plan_qty"]) else 0.0
                actual = float(row["actual_qty"]) if pd.notna(row.get("actual_qty")) else None

                data = {"product_id": pid, "yyyymm": ym, "plan_qty": plan, "actual_qty": actual}
                valid, errs = validate_sales_input(data)
                if valid:
                    insert_monthly_sales(pid, ym, plan, actual)
                    st.session_state[f"plan_{pid}"] = plan
                    success += 1
                else:
                    fail += 1
                    err_msgs.append(f"[{pid}] {', '.join(errs)}")

            if success:
                st.success(f"✅ {success}건 저장 완료")
            if fail:
                st.warning(f"⚠️ {fail}건 저장 실패")
                for m in err_msgs[:5]:
                    st.caption(m)

    except Exception as e:
        st.error(f"파일 처리 오류: {e}")
        import traceback
        st.code(traceback.format_exc())


def _render_leadtime_upload_section(products):
    """리드타임 실적 업로드 전용 섹션"""

    st.markdown("##### 📥 양식 다운로드")
    st.download_button(
        label="리드타임 입력양식 다운로드 (.xlsx)",
        data=_make_leadtime_template(products),
        file_name="리드타임실적_입력양식.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("##### 📤 파일 업로드")
    st.caption("필수 컬럼: `product_id` · `order_date` · `receipt_date` / `lt_days(일)` 비워두면 날짜 차이로 자동 계산")
    uploaded = st.file_uploader(
        "리드타임 실적 엑셀 파일을 여기에 올려주세요",
        type=["xlsx", "xls"],
        key="lt_file_uploader",
        label_visibility="collapsed",
    )

    if uploaded is None:
        st.info("양식을 다운받아 작성한 뒤 업로드하세요.")
        return

    try:
        xl = pd.ExcelFile(uploaded)
        # "리드타임실적" 시트가 있으면 우선 사용, 없으면 첫 번째 시트
        lt_sheets = [s for s in xl.sheet_names if "리드타임" in s or "leadtime" in s.lower()]
        sheet = lt_sheets[0] if lt_sheets else xl.sheet_names[0]
        df = xl.parse(sheet)
        df.columns = [c.replace("(일)", "").replace("(톤)", "").strip() for c in df.columns]

        required = {"product_id", "order_date", "receipt_date"}
        missing = required - set(df.columns)
        if missing:
            st.error(f"❌ 필수 컬럼 누락: {missing}  (읽은 시트: '{sheet}', 전체 시트: {xl.sheet_names})")
            return

        df["order_date"] = pd.to_datetime(df["order_date"])
        df["receipt_date"] = pd.to_datetime(df["receipt_date"])

        # lt_days 자동 계산
        if "lt_days" not in df.columns:
            df["lt_days"] = None
        mask = df["lt_days"].isnull() | (df["lt_days"].astype(str).str.strip() == "")
        df.loc[mask, "lt_days"] = (df.loc[mask, "receipt_date"] - df.loc[mask, "order_date"]).dt.days

        if "weight_qty" not in df.columns:
            df["weight_qty"] = None

        # 이상치 표시
        df["_outlier"] = df["lt_days"].apply(
            lambda x: "⚠️ 이상치" if pd.notna(x) and (float(x) < 0 or float(x) > 30) else "정상"
        )
        outlier_cnt = (df["_outlier"] == "⚠️ 이상치").sum()

        st.markdown("**📋 업로드 데이터 미리보기**")
        preview = df[["product_id", "order_date", "receipt_date", "lt_days", "weight_qty", "_outlier"]].copy()
        preview.columns = ["지종", "발주일", "입고일", "LT(일)", "중량(톤)", "상태"]
        st.dataframe(preview, use_container_width=True, hide_index=True)
        st.caption(f"총 {len(df)}행")

        if outlier_cnt:
            st.warning(f"⚠️ 이상치 {outlier_cnt}건 감지 (LT > 30일 또는 음수) → 절사평균으로 자동 보정 후 저장됩니다.")

        if st.button("💾 리드타임 실적 저장", type="primary", use_container_width=True, key="save_lt_btn"):
            success, fail, err_msgs = 0, 0, []
            for _, row in df.iterrows():
                pid = str(row["product_id"]).strip()
                order_d = str(row["order_date"].date())
                receipt_d = str(row["receipt_date"].date())
                lt_d = float(row["lt_days"]) if pd.notna(row["lt_days"]) else 0.0
                wt = float(row["weight_qty"]) if pd.notna(row.get("weight_qty")) else None

                data = {"product_id": pid, "order_date": order_d, "receipt_date": receipt_d, "lt_days": lt_d}
                valid, errs = validate_leadtime_input(data)
                if valid:
                    insert_leadtime_record(pid, order_d, receipt_d, lt_d, wt)
                    success += 1
                else:
                    fail += 1
                    err_msgs.append(f"[{pid}] {', '.join(errs)}")

            if success:
                st.success(f"✅ {success}건 저장 완료")
            if fail:
                st.warning(f"⚠️ {fail}건 저장 실패")
                for m in err_msgs[:5]:
                    st.caption(m)

    except Exception as e:
        st.error(f"파일 처리 오류: {e}")
        import traceback
        st.code(traceback.format_exc())


def render_input_page():
    st.title("📥 월별 실적 입력")

    products = get_products()
    target_yyyymm = _get_target_yyyymm()
    st.caption(f"산출 대상월: **{target_yyyymm[:4]}년 {target_yyyymm[4:]}월**")

    st.divider()

    # ── 업로드 2칸 나란히 ────────────────────────────────────────────────────
    col_sales, col_lt = st.columns(2, gap="large")

    with col_sales:
        st.subheader("📊 판매 실적 업로드")
        _render_sales_upload_section(products, target_yyyymm)

    with col_lt:
        st.subheader("🚚 리드타임 실적 업로드")
        _render_leadtime_upload_section(products)

    st.divider()

    # ── 운영재고 산출 트리거 ─────────────────────────────────────────────────
    st.subheader("🚀 운영재고 자동 산출")
    st.caption("입력한 대상월부터 해당 연도 12월까지 전체 월을 일괄 산출합니다. 미래 월 판매량은 수요예측으로 자동 추정됩니다.")
    target_input = st.text_input("산출 시작월 (YYYYMM)", value=target_yyyymm, max_chars=6)

    if st.button("▶ 운영재고 자동 산출 실행", type="primary", use_container_width=True):
        if not target_input or len(target_input) != 6 or not target_input.isdigit():
            st.error("산출 시작월을 YYYYMM 형식으로 입력하세요. (예: 202607)")
        else:
            end_month = f"{target_input[:4]}12"
            try:
                results, month_list = _run_full_pipeline(target_input, products)
                st.session_state["calc_results"] = results
                st.session_state["calc_target_yyyymm"] = target_input

                st.success(
                    f"✅ {target_input[:4]}년 {target_input[4:]}월 ~ 12월 "
                    f"({len(month_list)}개월) × {len(products)}개 지종 산출 완료!"
                )
                summary_df = pd.DataFrame([
                    {
                        "제품": r.get("product_name") or r["product_id"],
                        "안전재고(독립)": f"{r['safety_stock_independent']:,.1f}",
                        "사이클재고": f"{r['cycle_stock']:,.1f}",
                        "운영재고(톤)": f"{r['operating_stock']:,.1f}",
                        "운영재고(일)": f"{r['operating_days']:.1f}",
                        "플래그수": len(r.get("flags", [])),
                    }
                    for r in results
                ])
                st.dataframe(summary_df, use_container_width=True, hide_index=True)
                st.info(f"📊 '운영재고 산출 결과' 메뉴에서 {len(month_list)}개월 전체 추이를 확인하세요.")
            except Exception as e:
                st.error(f"산출 오류: {e}")
                import traceback
                st.code(traceback.format_exc())
