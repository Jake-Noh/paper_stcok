import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_results_to_excel(results_df, filename="운영재고_산출결과.xlsx"):
    """
    Create a formatted Excel workbook from results DataFrame.
    Returns bytes buffer suitable for st.download_button.
    """
    output = io.BytesIO()
    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "요약"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    warn_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    summary_cols = ["product_id", "target_yyyymm", "operating_stock", "operating_days",
                    "safety_stock_independent", "cycle_stock", "service_level"]
    col_labels = {
        "product_id": "제품코드",
        "target_yyyymm": "대상월",
        "operating_stock": "운영재고(톤)",
        "operating_days": "운영재고일수",
        "safety_stock_independent": "안전재고(독립)",
        "cycle_stock": "사이클재고",
        "service_level": "서비스수준",
    }

    # Write title
    ws_summary.merge_cells("A1:G1")
    title_cell = ws_summary["A1"]
    title_cell.value = "운영재고 산출 요약"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # Write headers
    for col_idx, col_key in enumerate(summary_cols, start=1):
        cell = ws_summary.cell(row=2, column=col_idx, value=col_labels.get(col_key, col_key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Write data rows
    df_summary = results_df[summary_cols] if all(c in results_df.columns for c in summary_cols) else results_df
    for row_idx, (_, row) in enumerate(df_summary.iterrows(), start=3):
        for col_idx, col_key in enumerate(summary_cols, start=1):
            val = row.get(col_key) if hasattr(row, "get") else row[col_key] if col_key in row.index else ""
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center")
            if col_key in ("operating_stock", "safety_stock_independent", "cycle_stock"):
                cell.number_format = "#,##0.0"
            elif col_key == "operating_days":
                cell.number_format = "0.0"
            elif col_key == "service_level":
                cell.number_format = "0.0%"

    # Totals row
    if len(df_summary) > 0:
        total_row = row_idx + 1
        ws_summary.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
        for col_idx, col_key in enumerate(summary_cols, start=1):
            if col_key in ("operating_stock", "safety_stock_independent", "cycle_stock"):
                col_letter = get_column_letter(col_idx)
                ws_summary.cell(
                    row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}3:{col_letter}{row_idx})"
                ).number_format = "#,##0.0"

    # Auto-width
    for col_idx in range(1, len(summary_cols) + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 16

    # ── Detail sheet ───────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("상세")
    if not results_df.empty:
        detail_headers = list(results_df.columns)
        for col_idx, col_name in enumerate(detail_headers, start=1):
            cell = ws_detail.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, (_, row) in enumerate(results_df.iterrows(), start=2):
            for col_idx, col_name in enumerate(detail_headers, start=1):
                val = row[col_name] if col_name in row.index else ""
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                # Flag warnings
                if col_name == "flags" and val and val != "[]" and val != "":
                    cell.fill = warn_fill

        for col_idx in range(1, len(detail_headers) + 1):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(output)
    output.seek(0)
    return output.read()
